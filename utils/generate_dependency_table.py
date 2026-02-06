#! /usr/bin/env python
"""
Script to generate table of PyHC package dependency conflicts.

__authors__ = ["Shawn Polson"]
"""


import openpyxl
from openpyxl.styles import PatternFill
from packaging.specifiers import SpecifierSet
from packaging.specifiers import Specifier
from packaging.version import Version
from packaging.version import InvalidVersion
import pandas as pd

from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import shlex
import subprocess

try:
    from utils.pipeline_utils import get_spec0_packages
except ModuleNotFoundError:
    from pipeline_utils import get_spec0_packages

# Named fills for spreadsheet highlighting
GREEN = PatternFill(start_color="00ff00", end_color="00ff00", fill_type="solid")
YELLOW = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
ORANGE = PatternFill(start_color="ffa500", end_color="ffa500", fill_type="solid")
RED = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
GRAY = PatternFill(start_color="aaaaaa", end_color="aaaaaa", fill_type="solid")
DARK_GRAY = PatternFill(start_color="333333", end_color="333333", fill_type="solid")


# TODO: get fisspy pipdeptree via conda (get-dep-tree-for-fisspy-w-conda.sh)
# TODO: sort final spreadsheet by lowercased names.

# TODO: Versions with wildcards * used to break everything.
#   Note: String comparison basically works if one version has * in it.
#   Original idea: (just removing * from versions for now; pretending like they don't exist) Can I just remove the * from the version when I encounter it? e.g. 5.0.* -> 5.0 (and still keep * in the eventual output?)
#   What I'm doing currently: Converting wildcard version specifiers to equivalent range specifiers in `remove_wildcards()`.

# TODO: Naming consideration: "project" = a (PyHC) package that gets its own table column,
#                             "package" = a (dependency) package that gets its own table row
#                             "dependency" = a combination of a "package" and a "version range"
#                             "version range" = a range like ">=1.5,2.0,!=1.6.*"
#                             vs
#                             "package" = any Python package
#                             "dependency" = a (dependency) package that gets its own table row
#                             "version range" = a range like ">=1.5,2.0,!=1.6.*"


def generate_requirements_file():
    package_names = get_core_pyhc_packages() + get_other_pyhc_packages()
    exclude = ["fisspy", "geopack", "ncarglow", "hwm93", "madrigalWeb", "maidenhead", "mgsutils", "pymap3d", "pysatCDF", "SkyWinder", "SkyWinder-Analysis", "TomograPy"]

    requirements = [name for name in package_names if name not in exclude]
    for package in requirements:
        print(package)
    return requirements


# def get_core_pyhc_packages():
#     """
#     TODO: consider scraping this from projects_core.yml online?
#     :return: A list of the core PyHC package names.

#     TODO:
#     kamodo has been removed until it supports Python>=3.12
#     """
#     return ["hapiclient", "plasmapy", "pysat", "pyspedas", "spacepy", "sunpy", "pyhc-core[tests]"]


# def get_other_pyhc_packages():
#     """
#     TODO: consider scraping this from projects.yml online? Would have to account for unpublished packages.
#     Note, the following packages are missing because they are not in PyPI:
#     ACEmag, AFINO, Auroral Electrojet (AEindex), fisspy, geodata, GIMAmag (gima-magnetometer), hwm93 (archived 2022),
#     iri90 (archived 2022), POLAN (archived 2022), MGSutils (archived 2022), ncarglow, PyGemini, pyglow, PyGS,
#     python-magnetosphere, sami2py, scanning-doppler-interferometer (archived 2022), TomograPy.

#     :return: A list of the other non-core PyHC package names.

#     TODO:
#     aidapy has been removed because it depends on heliopy which is incompatible with Python 3.12 and deprecated.
#     amisrsynthdata has been removed until they add support for Python 3.12.
#     heliopy has been removed due to incompatibility with Python 3.12 (it was originally hardcoded to 0.15.4 because 1.0.0 is deprecated)
#     pysatCDF has been removed due to installation failures.
#     """
#     return ["aacgmv2", "aiapy", "apexpy", "asilib", "astrometry-azel", "ccsdspy", "cdflib", "cloudcatalog", "dascutils", "dbprocessing", "dmsp", "enlilviz", "EUVpy", "fiasco", "gcmprocpy", "geopack", "georinex", "geospacelab", "goesutils", "hapiplot", "hissw", "igrf", "iri2016", "irispy-lmsal", "kaipy", "lofarSun", "lowtran", "madrigalWeb", "maidenhead", "mcalf", "msise00", "ndcube", "nexradutils", "ocbpy", "OMMBV", "pyaurorax", "pycdfpp", "pydarn", "pyflct", "pymap3d", "pyrfu", "pytplot", "pytplot-mpl-temp", "pyzenodo3", "reesaurora", "regularizepsf", "sammi-cdf", "savic", "sciencedates", "SciQLop==0.10.3", "SkyWinder", "solarmach", "solo-epd-loader", "space-packet-parser", "speasy", "spiceypy", "sunkit-image", "sunkit-instruments", "sunraster", "swxsoc", "themisasi", "viresclient", "wmm2015", "wmm2020"]


# def get_supplementary_packages():
#     """
#     :return: A list of supplementary packages, including optional dependencies not found by pipdeptree
#              and those used for unit tests.
#     TODO:
#     astropy=6.1.7 and xarray==2024.10.0 are hardcoded because the latest versions of the two are currently not compatible.
#     This was evidenced by a few PlasmaPy unit tests failing. Remove these pins when they become unnecessary.
#     """
#     return ["deepdiff", "hypothesis", "pytest-arraydiff", "pytest-doctestplus", "pytest-xdist", "setuptools-scm"] + ["astropy==6.1.7", "xarray==2024.10.0"]


# def spreadsheet_to_requirements_file_orig(spreadsheet):
#     # Extract PyHC package names (from the top row, excluding the first two columns)
#     # Remove '==' only if it appears at the end and is not followed by a version number
#     pyhc_packages = []
#     for pkg in spreadsheet.columns[2:]:
#         if pkg.endswith("=="):
#             pkg = pkg.replace("==", "")  # Remove '==' if no version number follows
#         pyhc_packages.append(pkg)
#
#     # Extract package names and their allowed version ranges
#     requirements = []
#     for index, row in spreadsheet.iterrows():
#         package = row['Package']
#         version_range = row['Allowed Version Range']
#         if pd.notna(version_range):
#             # Handle "Any" version specifier
#             if version_range.strip().lower() == 'any':
#                 requirement = package  # No version specifier for 'Any'
#             else:
#                 # Ensure no whitespace before '=='
#                 version_range = version_range.replace(" ", "")
#                 requirement = f"{package}{version_range}"
#             requirements.append(requirement)
#
#         # Remove duplicates: if a package from pyhc_packages is already in requirements, remove its version range
#         for pyhc_package in pyhc_packages:
#             # Extract package name from PyHC package, considering potential version specifier
#             pyhc_package_name = pyhc_package.split("==")[0] if "==" in pyhc_package else pyhc_package
#
#             # Find and remove any existing version of this package in requirements
#             requirements = [req for req in requirements if not req.startswith(pyhc_package_name)]
#
#     # Add PyHC packages to the requirements
#     requirements.extend(pyhc_packages)
#
#     # Alphabetize the list of requirements
#     requirements.sort(key=str.lower)
#
#     # Format as a requirements.txt content
#     requirements_txt = "\n".join(requirements)
#     return requirements_txt


def spreadsheet_to_requirements_file(spreadsheet_path):
    spreadsheet = pd.read_excel(spreadsheet_path, sheet_name=0)

    # Extract package names and their allowed version ranges
    requirements = []
    for index, row in spreadsheet.iterrows():
        package = row['Package']
        version_range = row['Allowed Version Range']

        # Check for "N/A" or NaN which indicates a conflict
        if pd.isna(version_range) or version_range.strip().lower() == 'n/a':
            raise ValueError(f"Cannot create a requirements file from the spreadsheet because there is a dependency conflict (with package `{package}`).")

        # Handle "Any" version specifier
        if version_range.strip().lower() == 'any':
            requirement = package  # No version specifier for 'Any'
        else:
            # Ensure no whitespace before version specifiers
            version_range = version_range.replace(" ", "")
            requirement = f"{package}{version_range}"
        requirements.append(requirement)

    # Alphabetize the list of requirements
    requirements.sort(key=str.lower)

    # Format as a requirements.txt content
    requirements_txt = "\n".join(requirements)
    return requirements_txt


def get_packages_installed_in_environment():
    """
    :return: A lowercased list of the package names output by `pip list`.
    """
    def parse_package_name(line):
        return line.strip().split(" ")[0].lower()

    pip_list_output = subprocess.check_output('pip list', shell=True).decode('utf-8')
    package_lines = pip_list_output.split("\n")[2:]
    return list(map(parse_package_name, package_lines))


def test_combine_ranges():
    r1 = combine_ranges(">=1.1.1", "any")
    r2 = combine_ranges("any", ">=1.1.2")
    r3 = combine_ranges(">=1.5", ">1.5,<2")
    r4 = combine_ranges(">1.1.1", "==1.2.3")
    # r5 = combine_ranges("<1.1.1", ">=1.2.3,<2.0")
    r6 = combine_ranges(">=1.1.1", ">=1.2.3,<2.0")
    r7 = combine_ranges(">=1.2.3,<2", ">=1.1.1")
    r8 = combine_ranges("==1.1.2", ">=0.9")
    r9 = combine_ranges("==1.1.2", "<2")
    r10= combine_ranges(">=1.9", ">=1.21.0")
    r11= combine_ranges(">=1.9", ">=1.19.5,<1.27.0")
    r12= combine_ranges(">=1.21.0", ">=1.19.5,<1.27.0")
    r13= combine_ranges(">=4.9.2", ">=4.12,!=5.0.*")  # TODO: how should I handle * wildcards?
    r14= combine_ranges(">=1.0.4", "==1.*")           # TODO: this should be allowed
    pass


def combine_ranges(current_range, new_range):
    """
    Combine the given ranges if they're compatible, otherwise raise a RuntimeError.
    :param current_range: String like ">=1.5.0,<2.0,!=1.6"
    :param new_range: String ">=1.5.0,<1.9"
    :return: String like ">=1.5.0,<1.9,!=1.6" or RuntimeError
    """
    if str(current_range).lower() == "any":
        return new_range
    if str(new_range).lower() == "any":
        return current_range

    temp_rules = SpecifierSet(current_range)
    new_rules = SpecifierSet(new_range)

    for new_rule in new_rules:
        if rule_is_compatible(temp_rules, new_rule):
            temp_rules = update_range(temp_rules, new_rule)
        else:
            raise RuntimeError(f"Found incompatibility: {current_range} vs. {new_range}")
    return str(temp_rules)


def update_range(current_range, new_rule):
    """
    Assuming new_rule is compatible with current_range, update current_range if new_rule requires it.
    Note: "any" rules are not supported (they should've been handled in `combine_ranges` by the time we reach here).
    :param current_range: SpecifierSet representing the currently-allowed version ranges for a given package.
    :param new_rule: A Specifier representing a new version requirement rule for the same package.
    :return: A SpecifierSet of the possibly-updated current_range.
    """
    op = new_rule.operator
    v = Version(new_rule.version)

    if op == "==":
        if v in current_range:
            return SpecifierSet(str(new_rule))  # current_range becomes '==v'
        else:
            return current_range
    elif op == "!=":
        # if we’re already pinned to a single explicit version, any extra
        # “!= …” rule is redundant, so ignore it.
        if any(spec.operator == "==" for spec in current_range):
            return current_range
        return update_exclusions(current_range, new_rule)  # Add exclusion '!=v' if it's not already there
    elif op == "~=":
        # Convert ~= to explicit bounds and merge regardless; if incompatible, it will be caught by
        # lower_bound_is_compatible/upper_bound_is_compatible via update_* calls
        v_parts = list(v.release)
        lower_bound = Specifier(f">={str(v)}")
        if len(v_parts) == 1:
            upper_parts = [v_parts[0] + 1]
        else:
            upper_parts = v_parts[:-1]
            upper_parts[-1] += 1
        upper_bound = Specifier(f"<{'.'.join(str(p) for p in upper_parts)}")
        temp_range = update_lower_bound(current_range, lower_bound)
        temp_range = update_upper_bound(temp_range, upper_bound)
        return temp_range
    elif op == ">" or op == ">=":
        if lower_bound_is_compatible(current_range, new_rule):
            return update_lower_bound(current_range, new_rule)  # new_rule might update/become current_range's lower bound
        else:
            return current_range
    elif op == "<" or op == "<=":
        if upper_bound_is_compatible(current_range, new_rule):
            return update_upper_bound(current_range, new_rule)  # new_rule might update/become current_range's upper bound
        else:
            return current_range


def rule_is_compatible(current_range, new_rule):
    """
    :param current_range: e.g. SpecifierSet(">=1.6,<2.0")
    :param new_rule: Specifier("==1.7")
    :return: Boolean for whether new_rule is compatible with current_range
    """
    op = new_rule.operator
    v = Version(new_rule.version)

    # TODO: clean up if/else/return logic?
    if op == "==":
        if v in current_range:
            return True  # POSSIBLE UPDATE: current_range becomes '==v'
        else:
            return False
    elif op == "!=":
        for spec in current_range:
            if spec == Specifier(f"=={v}"):
                return False
        return True  # POSSIBLE UPDATE: Add exclusion '!=v' (don't add a 2nd time if it's already in current_range)
    elif op == "~=":
        # Treat ~=X as the explicit window [X, bump(X)) and check both bounds against current_range
        # per PEP 440 compatible release behavior
        v_parts = list(v.release)
        if len(v_parts) == 1:
            upper_parts = [v_parts[0] + 1]
        else:
            upper_parts = v_parts[:-1]
            upper_parts[-1] += 1
        lower_bound = Specifier(f">={str(v)}")
        upper_bound = Specifier(f"<{'.'.join(str(p) for p in upper_parts)}")
        return lower_bound_is_compatible(current_range, lower_bound) and upper_bound_is_compatible(current_range, upper_bound)
    elif op == ">" or op == ">=":
        if lower_bound_is_compatible(current_range, new_rule):
            return True  # POSSIBLE UPDATE: new_rule might update/become current_range's lower bound
        else:
            return False
    elif op == "<" or op == "<=":
        if upper_bound_is_compatible(current_range, new_rule):
            return True  # POSSIBLE UPDATE: new_rule might update/become current_range's upper bound
        else:
            return False


def enforce_compatible_release(current_range, compatible_release_rule):
    """
    :param current_range: e.g. SpecifierSet(">=1.6,<2.0")
    :param compatible_release_rule: A Specifier compatible release rule e.g. Specifier("~=1.5")
    :return: A SpecifierSet that's the result of combining current_range with compatible_release_rule
    """
    if len(current_range) == 1 and str(current_range) == str(compatible_release_rule):
        return current_range  # they're both ~=v for the same v so no change

    # Convert ~= to explicit bounds
    v = Version(compatible_release_rule.version)
    v_parts = list(v.release)
    
    # Calculate the upper bound for ~=
    if len(v_parts) == 1:
        upper_parts = [v_parts[0] + 1]
    else:
        upper_parts = v_parts[:-1]
        upper_parts[-1] += 1
    
    # Create the explicit range
    lower_bound = Specifier(f">={str(v)}")
    upper_bound = Specifier(f"<{'.'.join(str(p) for p in upper_parts)}")
    
    # Combine with existing range
    temp_range = update_lower_bound(current_range, lower_bound)
    temp_range = update_upper_bound(temp_range, upper_bound)
    
    return temp_range


def update_upper_bound(current_range, upper_bound):
    """
    :param current_range: e.g. SpecifierSet(">=1.6,<2.0")
    :param upper_bound: A Specifier upper bound e.g. Specifier("<1.9")
    :return: A SpecifierSet of the possibly-updated current_range, e.g. SpecifierSet(">=1.6,<1.9")
    """
    rules = str(current_range).split(",")
    current_range_has_upper_bound = False
    for i, rule in enumerate(rules):
        if "<=" in rule:
            current_range_has_upper_bound = True
            if upper_bound.operator == "<" and Version(upper_bound.version) <= Version(Specifier(rule).version):
                rules[i] = str(upper_bound)
            elif upper_bound.operator == "<=" and Version(upper_bound.version) < Version(Specifier(rule).version):
                rules[i] = str(upper_bound)
        elif "<" in rule:
            current_range_has_upper_bound = True
            if Version(upper_bound.version) < Version(Specifier(rule).version):
                rules[i] = str(upper_bound)
        elif "==" in rule:
            return current_range  # '==version' will always trump upper_bound
        elif "~=" in rule:
            # If current_range contains a ~= rule, convert it to explicit bounds per PEP 440
            spec_rule = Specifier(rule)
            v = Version(spec_rule.version)
            v_parts = list(v.release)
            # Compute PEP 440 upper bound
            if len(v_parts) == 1:
                upper_parts = [v_parts[0] + 1]
            else:
                upper_parts = v_parts[:-1]
                upper_parts[-1] += 1
            v_lower = v
            v_upper = Version(".".join(str(p) for p in upper_parts))
            # Choose the strictest effective upper bound between incoming and ~= window cap
            effective_upper = Version(upper_bound.version)
            if effective_upper > v_upper:
                effective_upper = v_upper
            current_range_has_upper_bound = True
            # Replace the ~= rule with the effective upper bound
            rules[i] = str(Specifier(f"<{effective_upper}"))
            # Ensure the lower bound of the ~= window is present
            rules = str(update_lower_bound(SpecifierSet(",".join(rules)), Specifier(f">={v_lower}"))).split(",")
    if not current_range_has_upper_bound:
        rules.append(str(upper_bound))
    return SpecifierSet(",".join(rules))


def update_lower_bound(current_range, lower_bound):
    """
    TODO: handle when current_range is "!=5.0.*,>=4.9.2"?
    :param current_range: e.g. SpecifierSet(">=1.5,<2.0")
    :param lower_bound: A Specifier lower bound e.g. Specifier(">=1.6")
    :return: A SpecifierSet of the possibly-updated current_range, e.g. SpecifierSet(">=1.6,<2.0")
    """
    rules = str(current_range).split(",")
    current_range_has_lower_bound = False
    for i, rule in enumerate(rules):
        if ">=" in rule:
            current_range_has_lower_bound = True
            if lower_bound.operator == ">" and Version(lower_bound.version) >= Version(Specifier(rule).version):
                rules[i] = str(lower_bound)
            elif lower_bound.operator == ">=" and Version(lower_bound.version) > Version(Specifier(rule).version):
                rules[i] = str(lower_bound)
        elif ">" in rule:
            current_range_has_lower_bound = True
            if Version(lower_bound.version) > Version(Specifier(rule).version):
                rules[i] = str(lower_bound)
        elif "==" in rule:
            return current_range  # '==version' will always trump lower_bound
        elif "~=" in rule:
            # If current_range contains a ~= rule, convert it to explicit bounds per PEP 440
            spec_rule = Specifier(rule)
            v = Version(spec_rule.version)
            v_parts = list(v.release)
            # Compute PEP 440 upper bound
            if len(v_parts) == 1:
                upper_parts = [v_parts[0] + 1]
            else:
                upper_parts = v_parts[:-1]
                upper_parts[-1] += 1
            v_lower = v
            v_upper = Version(".".join(str(p) for p in upper_parts))
            if v_lower <= Version(lower_bound.version) < v_upper:
                current_range_has_lower_bound = True
                rules[i] = str(lower_bound)
                # Tighten/ensure the upper bound to be within ~= window if present
                rules = str(update_upper_bound(SpecifierSet(",".join(rules)), Specifier(f"<{v_upper}"))).split(",")
    if not current_range_has_lower_bound:
        rules.append(str(lower_bound))
    return SpecifierSet(",".join(rules))


def update_exclusions(current_range, exclusion):
    """
    :param current_range: e.g. SpecifierSet(">=1.6,<2.0")
    :param exclusion: Specifier exclusion e.g. Specifier("!=1.6.0")
    :return: SpecifierSet with exclusion added if it wasn't already in there
    """
    rules = str(current_range).split(",")
    if str(exclusion) in rules:
        return current_range  # exclusion is already in current_range
    else:
        return SpecifierSet(str(current_range) + "," + str(exclusion))  # add exclusion to current_range


def compatible_release_is_compatible(current_range, compatible_release_version):
    """
    :param current_range: e.g. SpecifierSet(">=1.6,<2.0")
    :param compatible_release_version: Version object from a compatible release rule (i.e. Version(v) from '~=v')
    :return: Boolean for whether compatible_release_version is compatible with current_range
    """
    v_list = list(compatible_release_version.release)
    v_list[-1] = "*"
    v_compatible = ".".join(str(e) for e in v_list)
    v_lower = Version(v_compatible.replace("*", "0"))
    v_upper = Version(v_compatible.replace("*", "999"))
    return v_lower in current_range or v_upper in current_range


def lower_bound_is_compatible(current_range, new_lower_bound):
    """
    :param current_range: e.g. SpecifierSet(">=1.6,<2.0")
    :param new_lower_bound: Specifier lower bound e.g. Specifier(">=1.5")
    :return: Boolean for whether new_lower_bound is compatible with current_range
    """
    inclusive = "=" in new_lower_bound.operator
    compatible = True
    for spec in current_range:
        curr_op = spec.operator
        curr_v = Version(spec.version)

        if curr_op == ">" or curr_op == ">=":
            pass
        elif curr_op == "<" or curr_op == "<=":
            if inclusive:
                compatible = Version(new_lower_bound.version) <= curr_v
            else:
                compatible = Version(new_lower_bound.version) < curr_v
            # Additional check to prevent empty intersection:
            if compatible and Version(new_lower_bound.version) == curr_v and new_lower_bound.operator.startswith('>=') and curr_op.startswith('<'):
                compatible = False
        elif curr_op == "==":
            compatible = (Version(new_lower_bound.version) <= curr_v and inclusive) or (Version(new_lower_bound.version) < curr_v)
        elif curr_op == "!=":
            pass
        elif curr_op == "~=":
            v_parts = list(curr_v.release)
            if len(v_parts) == 1:
                v_parts[0] += 1
            else:
                v_parts[-2] += 1
                v_parts = v_parts[:-1]
            v_upper = Version(".".join(str(p) for p in v_parts))
            compatible = Version(new_lower_bound.version) < v_upper

        if not compatible:
            return False
    return compatible


def upper_bound_is_compatible(current_range, new_upper_bound):
    """
    :param current_range: e.g. SpecifierSet(">=1.6,<2.0")
    :param new_upper_bound: Specifier upper bound e.g. Specifier("<1.9")
    :return: Boolean for whether new_upper_bound is compatible with current_range
    """
    inclusive = "=" in new_upper_bound.operator
    compatible = True
    for spec in current_range:
        curr_op = spec.operator
        curr_v = Version(spec.version)

        if curr_op == "<" or curr_op == "<=":
            pass
        elif curr_op == ">" or curr_op == ">=":
            if inclusive:
                compatible = Version(new_upper_bound.version) >= curr_v
            else:
                compatible = Version(new_upper_bound.version) > curr_v
            # Additional check for empty intersection:
            if compatible and Version(new_upper_bound.version) == curr_v and new_upper_bound.operator.startswith('<=') and curr_op.startswith('>'):
                compatible = False
        elif curr_op == "==":
            compatible = (Version(new_upper_bound.version) >= curr_v and inclusive) or (Version(new_upper_bound.version) > curr_v)
        elif curr_op == "!=":
            pass
        elif curr_op == "~=":
            # ~=3.0 means >=3.0,<4.0
            # Check if new upper bound doesn't exclude the lower part of the ~= range
            v_lower = curr_v  # The lower bound of ~= is the version itself
            compatible = Version(new_upper_bound.version) > v_lower

        if not compatible:
            return False
    return compatible


# def remove_wildcards(version_range_str):
#     """
#     Removes any * wildcards from the given versions by removing any occurrences of '.*' from their ends.
#     :param version_range_str: String like ">=1.5.0,<2.0,!=1.6.*"
#     :return: String like ">=1.5.0,<2.0,!=1.6"
#     """
#     def remove_wildcard(version_str):
#         release = version_str.split(".")
#         if release[-1] == "*":
#             release.pop()
#             return ".".join(release)
#         else:
#             # raise Exception("remove_wildcard() did not find a * at the end of '" + version_str + "' to remove.")
#             return version_str

#     versions = version_range_str.split(",")
#     return ",".join(list(map(remove_wildcard, versions)))

def remove_wildcards(version_range_str):
    """
    Converts wildcard version specifiers to equivalent range specifiers.
    :param version_range_str: String like ">=1.5.0,<2.0,!=1.6.*,==1.*"
    :return: String like ">=1.5.0,<2.0,!=1.6,>=1.0.0,<2.0.0"
    """
    if not version_range_str:
        return version_range_str
    
    # Split by comma to handle multiple specifiers
    specifiers = [spec.strip() for spec in version_range_str.split(",")]
    converted_specs = []
    
    for spec in specifiers:
        if not spec:
            continue
            
        # Extract operator and version
        if spec.startswith("=="):
            operator = "=="
            version = spec[2:].strip()
        elif spec.startswith("!="):
            operator = "!="
            version = spec[2:].strip()
        elif spec.startswith(">="):
            operator = ">="
            version = spec[2:].strip()
        elif spec.startswith("<="):
            operator = "<="
            version = spec[2:].strip()
        elif spec.startswith(">"):
            operator = ">"
            version = spec[1:].strip()
        elif spec.startswith("<"):
            operator = "<"
            version = spec[1:].strip()
        elif spec.startswith("~="):
            operator = "~="
            version = spec[2:].strip()
        else:
            # No operator, assume ==
            operator = "=="
            version = spec.strip()
        
        # Handle wildcard versions
        if version.endswith(".*"):
            base_version = version[:-2]  # Remove ".*"
            
            if operator == "==":
                # ==1.* becomes >=1.0.0,<2.0.0
                # ==1.2.* becomes >=1.2.0,<1.3.0
                parts = base_version.split(".")
                
                # Construct the lower bound (always pad with zeros to 3 parts)
                lower_parts = parts[:]
                while len(lower_parts) < 3:
                    lower_parts.append("0")
                lower_bound = ".".join(lower_parts)
                
                # Construct the upper bound (increment the LAST part of the base version)
                upper_parts = parts[:]
                try:
                    # Increment the last part in the original base version
                    last_idx = len(upper_parts) - 1
                    upper_parts[last_idx] = str(int(upper_parts[last_idx]) + 1)
                    
                    # Pad to 3 parts
                    while len(upper_parts) < 3:
                        upper_parts.append("0")
                    
                    upper_bound = ".".join(upper_parts)
                    
                    converted_specs.append(f">={lower_bound}")
                    converted_specs.append(f"<{upper_bound}")
                except ValueError:
                    # If we can't parse the version number, just remove the wildcard
                    converted_specs.append(f"{operator}{base_version}")
                    
            elif operator == "!=":
                # !=1.6.* is complex to represent exactly, but for practical purposes
                # we can approximate it. For now, just remove the wildcard and keep the !=
                # This is a limitation but better than the current behavior
                converted_specs.append(f"{operator}{base_version}")
                
            else:
                # For other operators, just remove the wildcard
                # This might not be perfect but is better than breaking
                converted_specs.append(f"{operator}{base_version}")
        else:
            # No wildcard, keep as-is
            converted_specs.append(spec)
    
    return ",".join(converted_specs)


def normalize_compatible_releases(version_range_str):
    """
    Expand any ~=X[.Y[.Z]] specifiers into explicit >=X[.Y[.Z]] and <bump(X[.Y]).
    Leaves other specifiers unchanged.
    """
    if not version_range_str:
        return version_range_str
    parts = [p.strip() for p in version_range_str.split(",") if p.strip()]
    out = []
    for p in parts:
        if p.startswith("~="):
            ver = p[2:].strip()
            v = Version(ver)
            r = list(v.release)
            if len(r) == 1:
                upper = [r[0] + 1]
            else:
                upper = r[:-1]
                upper[-1] += 1
            out.append(f">={v}")
            out.append(f"<{'.'.join(str(x) for x in upper)}")
        else:
            out.append(p)
    return ",".join(out)

# Test the improved function
def test_improved_remove_wildcards():
    """Test cases for the improved remove_wildcards function"""
    test_cases = [
        ("==1.*", ">=1.0.0,<2.0.0"),
        (">=1.0.4,==1.*", ">=1.0.4,>=1.0.0,<2.0.0"), 
        ("!=1.6.*", "!=1.6"),  # Approximation for now
        (">=4.9.2,!=5.0.*", ">=4.9.2,!=5.0"),
        ("==1.2.*", ">=1.2.0,<1.3.0"),
        (">=1.5.0,<2.0,!=1.6.*", ">=1.5.0,<2.0,!=1.6"),
        (">=1.0.0", ">=1.0.0"),  # No wildcards
        ("", ""),  # Empty string
    ]
    
    print("Testing improved remove_wildcards function:")
    for input_val, expected in test_cases:
        result = remove_wildcards(input_val)
        print(f"Input:    {input_val}")
        print(f"Expected: {expected}")
        print(f"Got:      {result}")
        print(f"Match:    {result == expected}")
        print("---")


def determine_version_range(dependencies, package, requirement):
    """
    :param dependencies: Dict like {'package1': '>=1.0', 'package2': '<23.0', ...}
    :param package: String name of a package like 'package1'
    :param requirement: String range like ">=1.5,<2.0"
    :return: A String range like ">=1.5,<2.0" that's either requirement or the (possibly-updated) dependencies[package]
    """
    if package not in dependencies:
        return requirement
    else:
        current_range = dependencies[package]
        try:
            new_range = reorder_requirements(combine_ranges(current_range, requirement))
            return new_range
        except RuntimeError as e:
            raise RuntimeError(f"Package '{package}': {str(e)}")


def get_base_package_name(package):
    """
    Extract the base package name from a package specifier.
    Examples:
      - "sunpy==7.1.0" -> "sunpy"
      - "pyhc-core[tests]==0.0.7" -> "pyhc-core"
    """
    package = package.strip()
    return re.split(r"[=<>!\[\s]", package)[0]


def parse_uv_tree_output(package, output_str):
    """
    Parse `uv pip tree --show-version-specifiers` output.
    Returns:
      (package_version, dependencies)
    where dependencies is Dict like {"numpy": ">=2.0,<3.0", ...}
    """
    lines = [line.strip() for line in output_str.splitlines() if line.strip()]
    if not lines:
        raise RuntimeError(
            f"Failed to parse dependency tree for package '{package}': empty output."
        )

    root_match = re.match(r"^([A-Za-z0-9_.-]+)\s+v([^\s]+)$", lines[0])
    if not root_match:
        raise RuntimeError(
            f"Failed to parse version for package '{package}'. "
            f"The first line of uv pip tree output did not match '<name> v<version>'.\n"
            f"Output:\n{output_str}"
        )
    package_version = root_match.group(2)

    dependencies = {}
    dep_pattern = re.compile(
        r"^[│\s├└─]*([A-Za-z0-9_.-]+)\s+v[^\s]+\s+\[(required|requires):\s+(.+?)\]\s*$"
    )
    for line in lines[1:]:
        match = dep_pattern.match(line)
        if not match:
            continue
        name = match.group(1).lower()
        version_range = match.group(3).strip()
        if version_range == "*":
            version_range = "any"
        else:
            version_range = ",".join(part.strip() for part in version_range.split(","))
        version_range = remove_wildcards(version_range)
        version_range = normalize_compatible_releases(version_range)
        dependencies[name] = determine_version_range(dependencies, name, version_range)
    return package_version, dependencies


def _build_dependency_tree_command(package, use_installed, installed_packages):
    base_package = get_base_package_name(package)
    base_package_lower = base_package.lower()

    if use_installed and base_package_lower in installed_packages:
        return f"uv pip tree --show-version-specifiers --package {shlex.quote(base_package)}"

    if package.startswith("git+"):
        return f"./utils/get-dep-tree-for-git-package.sh {shlex.quote(package)}"
    if base_package == "pysatCDF":
        return f"./utils/get-dep-tree-for-pysatCDF-w-numpy.sh {shlex.quote(package)}"
    # if base_package == "OMMBV":
    #     return f"./utils/get-dep-tree-for-ommbv.sh {shlex.quote(package)}"
    # if base_package == "spacepy":
    #     return f"./utils/get-dep-tree-for-spacepy.sh {shlex.quote(package)}"
    # if base_package == "fisspy":
    #     return f"./get-dep-tree-for-fisspy-w-conda.sh {shlex.quote(package)}"
    if base_package in {"asilib", "pyaurorax"}:
        return f"./utils/get-dep-tree-for-package-w-opencv-python.sh {shlex.quote(package)}"
    if base_package in {"cloudcatalog", "pyrfu", "swxsoc"}:
        return f"./utils/get-dep-tree-for-package-w-boto.sh {shlex.quote(package)}"
    if base_package in {"EUVpy", "kaipy", "SciQLop"}:
        return f"./utils/get-dep-tree-for-package-w-httpcore.sh {shlex.quote(package)}"
    return f"./utils/get-dep-tree-for-package.sh {shlex.quote(package)}"


def _get_package_dependencies(package, use_installed, installed_packages):
    command = _build_dependency_tree_command(package, use_installed, installed_packages)
    output_str = subprocess.check_output(command, shell=True, text=True)
    package_version, dependencies = parse_uv_tree_output(package, output_str)
    dependencies[get_base_package_name(package)] = f"=={package_version}"
    sorted_dependencies = {key: value for key, value in sorted(dependencies.items())}
    package_w_version = f"{package}=={package_version}" if "==" not in package else package
    return package_w_version, sorted_dependencies


def get_dependency_ranges_by_package(packages, use_installed=False, max_workers=1):
    """
    TODO: rename func to "get_dependency_ranges/requirements_for_packages()"?
    TODO: go back to "by project" wording?
    Gets each package's dependency requirements by creating temporary python environments to ensure pip installs work.
    Pre-installed package versions get used when use_installed is True, otherwise the latest package versions get used.
    :param packages: List of packages like ['hapiclient', 'sunpy'] that may not be compatible together
    :param use_installed A Boolean for whether to try to use pre-installed package versions
    :param max_workers: Number of worker threads to use when extracting package trees.
    :return: Dict like {'hapiclient': {'package1': '>=1.0'}, 'sunpy': {...}} (dependencies sorted alphabetically)
    """
    if max_workers < 1:
        max_workers = 1

    installed_packages = set(get_packages_installed_in_environment())
    all_dependencies = {}
    total_packages = len(packages)  # for progress tracking output

    def _process_single_package(index, package):
        print(
            f"Processing package: {package} ({index}/{total_packages})",
            flush=True,
        )
        package_w_version, dependencies = _get_package_dependencies(
            package, use_installed, installed_packages
        )
        return index, package_w_version, dependencies

    indexed_packages = list(enumerate(packages, start=1))
    if max_workers == 1:
        for index, package in indexed_packages:
            _, package_w_version, dependencies = _process_single_package(index, package)
            all_dependencies[package_w_version] = dependencies
        return all_dependencies

    ordered_results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_package = {
            executor.submit(_process_single_package, index, package): (index, package)
            for index, package in indexed_packages
        }
        for future in as_completed(future_to_package):
            index, package = future_to_package[future]
            try:
                _, package_w_version, dependencies = future.result()
            except Exception as e:
                raise RuntimeError(f"Failed to process package '{package}': {e}") from e
            ordered_results[index] = (package_w_version, dependencies)

    for index in sorted(ordered_results):
        package_w_version, dependencies = ordered_results[index]
        all_dependencies[package_w_version] = dependencies
    return all_dependencies


def get_dependency_ranges_for_environment(env_packages, full_path_to_pipdeptree=None):
    """
    TODO: rename func to "find_common_environment_dependency_ranges/requirements()"
    :param env_packages: List of packages like ['hapiclient', 'sunpy'] expected to already be installed in the env
    :param full_path_to_pipdeptree: Optional command override. If provided, this should point to a
                                    `uv pip tree --show-version-specifiers` compatible command prefix.
    :return: Dict like {'package1': '>=1.0', 'package2': '<23.0', ...} (sorted alphabetically)
    """
    uv_tree_command = (
        full_path_to_pipdeptree
        if full_path_to_pipdeptree
        else "uv pip tree --show-version-specifiers"
    )
    dependencies = {}
    for package in env_packages:
        base_package = get_base_package_name(package)
        command = f"{uv_tree_command} --package {shlex.quote(base_package)}"
        output_str = subprocess.check_output(command, shell=True, text=True)
        _, package_dependencies = parse_uv_tree_output(package, output_str)
        for name, version_range in package_dependencies.items():
            dependencies[name] = determine_version_range(dependencies, name, version_range)
    sorted_dependencies = {k: v for k, v in sorted(dependencies.items())}
    return sorted_dependencies


def reduce_environment_requirements(project_dependencies, allow_conflicts=True):
    """
    :param project_dependencies: Dict like {'PyHC Package 1': {'package1': '>=1.0'}, 'PyHC Package 2': {...}}.
    :param allow_conflicts: Boolean that, when False, will raise an exception the moment a conflict is found.
    :return: Dict like {'package1': '>=1.0,<2.0', 'package2': 'None', ...} made by combining all package requirements (sorted alphabetically).
             Note: if dependency conflicts exist for a package and allow_conflicts=True, its allowed range will be None.
    """
    all_dependencies = {}
    for project, project_dependencies in project_dependencies.items():
        for package_name, v_range in project_dependencies.items():
            if package_name in all_dependencies:
                current_range = all_dependencies[package_name]
                if current_range is None:
                    pass  # at least one dependency conflict exists for this package so no valid range exists
                else:
                    if are_compatible(current_range, v_range):
                        all_dependencies[package_name] = reorder_requirements(combine_ranges(current_range, v_range))
                    else:
                        if allow_conflicts:
                            all_dependencies[package_name] = None  # found a dependency conflict for this package
                        else:
                            raise RuntimeError(f"Found conflict in '{project}': {package_name} {v_range}")
            else:
                all_dependencies[package_name] = v_range
    sorted_dependencies = {key: value for key, value in sorted(all_dependencies.items())}
    return sorted_dependencies


def are_compatible(allowed_range, package_range):
    """
    :param allowed_range: String like ">=1.5.0,<2.0,!=1.6"
    :param package_range: String ">=1.5.0,<1.9"
    :return: Boolean of whether package_range is compatible with allowed_range
    """
    try:
        r = combine_ranges(allowed_range, package_range)
        return True
    except RuntimeError:
        return False


def is_spec0_compliant(package_name, version_range, spec0_requirements):
    """
    Check if a package's version range is compatible with SPEC 0 requirements.

    :param package_name: String like "numpy"
    :param version_range: String like ">=1.5.0,<2.0" or None
    :param spec0_requirements: Dict like {"numpy": ">=2.0.0", "scipy": ">=1.12.0", ...}
    :return: Boolean or None (None if package not in SPEC 0)
    """
    spec0_key = package_name.lower()
    if spec0_key not in spec0_requirements:
        return None  # Not a SPEC 0 package

    if version_range is None or not isinstance(version_range, str):
        return False

    cleaned_range = version_range.strip()
    if not cleaned_range or cleaned_range.lower() == "none":
        return False  # No versions allowed

    if cleaned_range.lower() == "any":
        return True  # Any version includes SPEC 0 compliant versions

    # Normalize range to handle wildcards, ~=, and other quirky specs
    normalized_range = normalize_compatible_releases(remove_wildcards(cleaned_range))
    normalized_range = reorder_requirements(normalized_range)
    if not normalized_range:
        return False

    # Check if ranges are compatible (SPEC 0 requirement first, project range second)
    spec0_specifier = spec0_requirements[spec0_key]
    return are_compatible(spec0_specifier, normalized_range)


def _format_requirement_for_text(package_name, version_range):
    """
    Format a package requirement like "numpy>=2.0.0" for text output.
    """
    if version_range is None:
        return package_name
    cleaned = str(version_range).strip()
    if not cleaned:
        return package_name
    return f"{package_name}{cleaned}"


def find_spec0_problems(table_data):
    """
    Return a list of human-readable SPEC 0 incompatibilities found in table_data.
    """
    if not isinstance(table_data, dict):
        return []
    project_data = table_data.get("project_data")
    spec0_requirements = table_data.get("spec0_requirements", {})
    if not isinstance(project_data, dict) or not spec0_requirements:
        return []

    problems = []
    for project, dependency_data in project_data.items():
        project_name = project.split("==")[0] if isinstance(project, str) else str(project)
        for package_name, values in dependency_data.items():
            if not isinstance(values, tuple) or len(values) < 3:
                continue
            _, spec0_compliant, version_range = values
            if spec0_compliant is False:
                spec0_req = spec0_requirements.get(package_name.lower())
                if not spec0_req:
                    continue
                project_req = _format_requirement_for_text(package_name, version_range)
                spec0_req_text = _format_requirement_for_text(package_name, spec0_req)
                problems.append((project_name, package_name, project_req, spec0_req_text))

    problems = sorted(problems, key=lambda x: (x[0].lower(), x[1].lower()))
    return [f"{project}: requires {project_req} but SPEC 0 requires {spec0_req}"
            for project, _, project_req, spec0_req in problems]


def clean_dependencies(dependencies):
    """
    Given a dependencies dict, add a space before version ranges that start with '=',
    and reorder version ranges as: lower bound, upper bound, then exclusions/remaining.
    :param dependencies: Dict like {'package1': '<2,>=1.0', 'package2': '==23.0', ...}
    :return: Dict like {'package1': '>=1.0,<2', 'package2': ' ==23.0', ...}
    """
    if dependencies:
        cleaned_deps = {}
        for package, range_str in dependencies.items():
            range_str = clean_range_str(range_str)
            cleaned_deps[package] = range_str
        return cleaned_deps
    else:
        return dependencies


def clean_range_str(range_str):
    """
    Given a version range string, add a space before version ranges that start with '=',
    and reorder version ranges as: lower bound, upper bound, then exclusions/remaining.
    :param range_str: String like "<2.0,!=1.6,>=1.5.0"
    :return: Cleaned version range string like ">=1.5.0,<2.0,!=1.6"
    """
    if range_str:
        range_str = reorder_requirements(range_str)
        if range_str.startswith("="):
            range_str = " " + range_str
    return range_str


def reorder_requirements(range_str):
    """
    Reorder version ranges: lower bound, upper bound, then exclusions.
    :param range_str: String like "<2.0,!=1.6,>=1.5.0,"
    :return: Reordered String like ">=1.5.0,<2.0,!=1.6"
    """
    if range_str:
        rules = [rule.strip() for rule in range_str.split(",") if rule.strip()]
        lower_bound = None
        upper_bound = None
        remaining = []
        for rule in rules:
            if rule.startswith(">"):
                lower_bound = rule
            elif rule.startswith("<"):
                upper_bound = rule
            else:
                remaining.append(rule)

        reordered_range_str = ""
        if lower_bound:
            reordered_range_str += lower_bound
        if upper_bound:
            reordered_range_str += upper_bound if not reordered_range_str else f',{upper_bound}'
        if remaining:
            reordered_range_str += ",".join(remaining) if not reordered_range_str else f',{",".join(remaining)}'
        return reordered_range_str
    else:
        return range_str


def compare_requirements(env_dependencies, package_dependencies):  # TODO: delete me?
    """
    TODO: Should probably delete this func...
    TODO: handle "found incompatibility" exceptions: Idea: try/catch, rather than let "Found incompatibility" string become current_range, store the incompatibility in a list of them such that all_deps[package] = tuple(deps, incompatible_rules)? Well... actually, I think that'd have to happen in `are_compatible()` because these ranges are just for individual projects so there won't be conflicts... `are_compatible()` is where we compare a project's range to the env's allowed range...
    :param env_dependencies: Dict like {'package1': '>=1.0', 'package2': '<23.0', ...}
    :param package_dependencies: Dict like {'PyHC Package 1': {'package1': '>=1.0'}, 'PyHC Package 2': {...}}
    :return compatibility: Dict like {'PyHC Package 1': {'package1': (True,'>=1.0'), 'package2': (False,'<23.0')}, ...}
             TODO: {'Env Dependencies': env_dependencies, 'Package Dependencies': compatibility}
    """
    compatibility = {}
    for package, package_deps in package_dependencies.items():
        package_deps = clean_dependencies(package_deps)
        compatibility[package] = {}
        for dep, allowed_range in env_dependencies.items():  # TODO: BUG: this misses any new dependencies in the non-core packages!! Need to loop through `package_deps` instead somehow...
            if dep in package_deps:
                package_range = package_deps[dep]
                package_is_compatible = are_compatible(allowed_range, package_range)
                compatibility[package][dep] = (package_is_compatible, package_range)
            else:
                compatibility[package][dep] = (None, None, None)
    return compatibility


def generate_first_two_cols_of_dependency_spreadsheet(env_dependencies):  # TODO: delete me...?
    """
    :param env_dependencies: Dict like {'package1': '>=1.0', 'package2': '<23.0', ...}
    :return: Excel workbook of the first two columns of the dependency conflict table
    """
    if not isinstance(env_dependencies, dict) or not env_dependencies:
        raise ValueError('Invalid dependencies input')

    # Prepare data for Excel
    env_dependencies = clean_dependencies(env_dependencies)

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # Write headers to the worksheet
    worksheet['A1'] = 'Package'
    worksheet['B1'] = 'Allowed Version Range'

    # Write data to the worksheet
    for package, version_range in env_dependencies.items():
        row = [package, version_range]
        worksheet.append(row)

    return workbook


def generate_dependency_spreadsheet(env_dependencies, package_dependencies):  # TODO: delete me...?
    """
    :param env_dependencies: Dict like {'package1': '>=1.0', 'package2': '<23.0', ...}
    :param package_dependencies: Dict like {'PyHC Package 1': {'package1': '>=1.0'}, 'PyHC Package 2': {...}}
    :return: Excel workbook of the dependency conflict table
    """
    if not isinstance(env_dependencies, dict) or not env_dependencies:
        raise ValueError('Invalid env_dependencies input')
    if not isinstance(package_dependencies, dict) or not package_dependencies:
        raise ValueError('Invalid package_dependencies input')

    # Create a new workbook populated with our environment's dependency requirements
    workbook = generate_first_two_cols_of_dependency_spreadsheet(env_dependencies)

    # Use a data structure of compatibilities to fill remaining columns for our package dependencies
    compatibility = compare_requirements(env_dependencies, package_dependencies)

    # Get the worksheet from our workbook
    worksheet = workbook[workbook.sheetnames[0]]

    # Add a column for each PyHC package to the worksheet
    for i, project in enumerate(compatibility.keys(), start=3):  # TODO: use "package" rather than "project"?
        worksheet.cell(row=1, column=i, value=project)

        # Check each row in the column for compatibility with the project's dependency requirements
        for j, (dep, (dep_compatibility, dep_range)) in enumerate(compatibility[project].items(), start=2):
            cell = worksheet.cell(row=j, column=i)
            if dep_compatibility is not None:
                cell.value = dep_range
                if dep_compatibility:
                    cell.fill = PatternFill(start_color="00ff00", end_color="00ff00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")

    return workbook


def merge_dependencies(core_dependencies, other_dependencies):
    """
    Merges two dependency dicts such that non-conflicting keys in other_dependencies are directly appended
    to core_dependencies (in order).
    :param core_dependencies: Dict of core dependencies like {'package1': '>=1.0', 'package2': '<23.0', ...}
    :param other_dependencies: Dict of other dependencies like {'package2': '>=24.0', 'package3': '<1.0', ...}
    :return: Combined dict like {'package1': '>=1.0', 'package2': '<23.0', 'package3': '<1.0', ...}
    """
    merged_dict = {**core_dependencies, **other_dependencies}
    for key, value in merged_dict.items():
        if key in core_dependencies:
            merged_dict[key] = core_dependencies[key]
    return merged_dict


# TODO: Need func table_data_to_2D_array(table_data) that flattens table_data
#       (biggest change: lots of {'package': (None, None, None)} cells where projects don't use that dependency).


def generate_dependency_table_data(packages, core_env_packages=[], max_workers=1):
    """
    Generates a data structure that can populate a dependency conflict table.
    :param packages: A list of PyHC packages ["package1", "package2", ...] that may have dependency conflicts
    :param core_env_packages: A list of PyHC packages ["package1", "package2", ...] that DON'T have dependency conflicts (assumed to already be installed in env)
    :return: A dict like {
                          'core_dependencies':
                              {'package1': (2, '>=1.0'), 'package2': (3, '<23.0'), ...},
                          'other_dependencies':
                              {'package3': (5, 'None'), 'package4': (6, '>=1.5,<2'), ...},
                          'project_data':
                              {
                               'PyHC Package 1':
                                  {'package1': (True,'>=1.0'), 'package2': (False,'<23.0')},
                               'PyHC Package 2': ...
                              }
                         }
             with version range data cleaned for Excel.
    """
    core_deps_by_project = get_dependency_ranges_by_package(
        core_env_packages,
        use_installed=True,
        max_workers=max_workers,
    )
    other_deps_by_project = get_dependency_ranges_by_package(
        packages,
        max_workers=max_workers,
    )
    all_deps_by_project = {**core_deps_by_project, **other_deps_by_project}

    # import pickle  # TODO: delete pickling
    # with open('core_deps_by_project3.pkl', 'wb') as f:
    #     pickle.dump(core_deps_by_project, f)
    # with open('other_deps_by_project3.pkl', 'wb') as f:
    #     pickle.dump(other_deps_by_project, f)
    # with open('all_deps_by_project3.pkl', 'wb') as f:
    #     pickle.dump(all_deps_by_project, f)

    # import pickle
    # with open('core_deps_by_project8.pkl', 'rb') as f:
    #     core_deps_by_project = pickle.load(f)
    # with open('other_deps_by_project8.pkl', 'rb') as f:
    #     other_deps_by_project = pickle.load(f)
    # # with open('all_deps_by_project8.pkl', 'rb') as f:
    # #     all_deps_by_project = pickle.load(f)
    # all_deps_by_project = {**core_deps_by_project, **other_deps_by_project}
    #
    # other_deps_by_project['enlilviz']['numpy'] = "<1.0.0"
    # other_deps_by_project['geospacelab']['pandas'] = "<1.0.0"

    core_dependencies = reduce_environment_requirements(core_deps_by_project, allow_conflicts=False)
    other_dependencies = reduce_environment_requirements(other_deps_by_project)
    other_dependencies = {k: v for k, v in other_dependencies.items() if k not in core_dependencies}  # make unique
    all_dependencies = merge_dependencies(core_dependencies, other_dependencies)

    # Parse SPEC 0 requirements once
    spec0_reqs_list = get_spec0_packages()  # ["numpy>=2.0.0", "scipy>=1.11.0", ...]
    spec0_requirements = {}
    for req in spec0_reqs_list:
        # Parse "numpy>=2.0.0" into {"numpy": ">=2.0.0"}
        match = re.match(r'^([a-zA-Z0-9_-]+)(.*)$', req)
        if match:
            pkg_name = match.group(1).lower()
            version_spec = match.group(2).strip()
            if version_spec:
                spec0_requirements[pkg_name] = version_spec

    table_data = {
        'core_dependencies': core_dependencies,
        'other_dependencies': other_dependencies,
        'project_data': {},
        'spec0_requirements': spec0_requirements
    }
    for project, project_dependencies in all_deps_by_project.items():
        table_data['project_data'][project] = {}
        for package_name, version_range in project_dependencies.items():
            allowed_range = all_dependencies[package_name]
            # Always check SPEC 0 compliance for the individual package's version range
            spec0_compliant = is_spec0_compliant(package_name, version_range, spec0_requirements)
            if allowed_range is not None:
                compatible = are_compatible(allowed_range, version_range)
                table_data['project_data'][project][package_name] = (compatible, spec0_compliant, clean_range_str(version_range))
            else:
                # Conflict: allowed_range is None, but we still check SPEC 0 compliance
                table_data['project_data'][project][package_name] = (False, spec0_compliant, clean_range_str(version_range))

    # Clean data and add row numbers (and SPEC 0 compliance for "Allowed Version Range" column)
    core_dependencies = clean_dependencies(table_data['core_dependencies'])
    core_dependencies = {k: (i, v, is_spec0_compliant(k, v, spec0_requirements)) for i, (k, v) in enumerate(core_dependencies.items(), start=2)}
    table_data['core_dependencies'] = core_dependencies

    start = len(table_data['core_dependencies']) + 3 if core_dependencies else 2
    other_dependencies = clean_dependencies(table_data['other_dependencies'])
    other_dependencies = {k: (i, v, is_spec0_compliant(k, v, spec0_requirements)) for i, (k, v) in enumerate(other_dependencies.items(), start=start)}
    table_data['other_dependencies'] = other_dependencies

    return table_data


def pad_table_project_data(table_data):
    """
    TODO: consider using collections.OrderedDict if order becomes a problem.
    TODO: put a bunk "N/A" -> (None, None, None) <unlikely-separator-val> entry in project_data between core and other deps, if there are core deps
    :param table_data: The data structure returned from `generate_dependency_table_data()`
    :return: table_data but with its project_data padded with None data for dependencies that projects don't use.
    """
    core_dependencies = table_data['core_dependencies']
    other_dependencies = table_data['other_dependencies']
    project_data = table_data['project_data']
    temp_data = project_data

    for project, dependency_data in project_data.items():
        if core_dependencies:
            # pad cells for core_dependencies,
            # add {"N/A": (None, None, None)} as separator
            for package_name in core_dependencies.keys():
                if package_name not in dependency_data:
                    temp_data[project][package_name] = (None, None, None)
            sorted_data = {k: v for k, v in sorted(temp_data[project].items())}
            temp_data[project] = sorted_data
            temp_data[project]["N/A"] = (None, None, None)  # separator between core and other dependencies

        # pad for other_dependencies
        other_temp_data = {}
        for package_name in other_dependencies.keys():
            if package_name not in dependency_data:
                #temp_data[project][package_name] = (None, None, None)
                other_temp_data[package_name] = (None, None, None)
            else:
                other_temp_data[package_name] = dependency_data[package_name]

        temp_data[project] = merge_dependencies(temp_data[project], other_temp_data)

    table_data['project_data'] = temp_data
    return table_data


# def excel_spreadsheet_from_table_data(table_data):
#     """
#     :param table_data: The data structure returned from `generate_dependency_table_data()`
#     :return: The Excel workbook that is the dependency conflict table
#     """
#     if not isinstance(table_data, dict) or not table_data:
#         raise ValueError('Invalid table data.')
#     # TODO: assert: len of project data should be the length of core + other dependencies +1 (when there are core ones)
#
#     # Pad data for Excel extraction
#     # table_data = pad_table_project_data(table_data)  # TODO: consider renaming funcs because of this (work it into `generate_dependency_table_data()`?
#
#     # Extract data for Excel
#     core_dependencies = table_data['core_dependencies']
#     other_dependencies = table_data['other_dependencies']
#     project_data = table_data['project_data']
#
#     # Create a new workbook
#     workbook = openpyxl.Workbook()
#
#     # Select the active worksheet
#     worksheet = workbook.active
#
#     # Write headers to the worksheet
#     worksheet['A1'] = 'Package'
#     worksheet['B1'] = 'Allowed Version Range'
#
#     # Write data to the worksheet
#     if core_dependencies:
#         # Write core environment dependencies to the first two columns of the worksheet (shaded light gray)
#         for i, (package_name, version_range) in enumerate(core_dependencies.items(), start=2):
#             # Column 1 - Package
#             cell = worksheet.cell(row=i, column=1)
#             cell.value = package_name
#             cell.fill = PatternFill(start_color="aaaaaa", end_color="aaaaaa", fill_type="solid")
#
#             # Column 2 - Allowed Version Range
#             cell = worksheet.cell(row=i, column=2)
#             cell.value = version_range  # TODO: shouldn't ever be "N/A" but should I consider that case here anyway?
#             cell.fill = PatternFill(start_color="aaaaaa", end_color="aaaaaa", fill_type="solid")
#
#         # Add a dark gray separator before other dependencies
#         cell = worksheet.cell(row=i+1, column=1)
#         cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
#         cell = worksheet.cell(row=i+1, column=2)
#         cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
#
#     # Write other environment dependencies to the first two columns of the worksheet
#     for package_name, version_range in other_dependencies.items():
#         row = [package_name, version_range if version_range is not None else "N/A"]
#         worksheet.append(row)
#
#     # Write project data to the worksheet one column at a time
#     for j, (project, dependency_data) in enumerate(project_data.items(), start=3):
#         worksheet.cell(row=1, column=j, value=project)  # column header
#         for i, (package_name, (compatible, version_range)) in enumerate(dependency_data.items(), start=2):
#             cell = worksheet.cell(row=i, column=j)
#             if compatible is not None:
#                 cell.value = version_range
#                 if compatible:
#                     cell.fill = PatternFill(start_color="00ff00", end_color="00ff00", fill_type="solid")
#                 else:
#                     cell.fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
#
#     return workbook


def excel_spreadsheet_from_table_data(table_data):
    """
    :param table_data: The data structure returned from `generate_dependency_table_data()`
    :return: The Excel workbook that is the dependency conflict table
    """
    if not isinstance(table_data, dict) or not table_data:
        raise ValueError('Invalid table data.')
    if not isinstance(table_data['core_dependencies'], dict):
        raise ValueError('Invalid core_dependencies in table data.')
    if not isinstance(table_data['other_dependencies'], dict):
        raise ValueError('Invalid other_dependencies in table data.')
    if not isinstance(table_data['project_data'], dict):
        raise ValueError('Invalid project_data in table data.')

    # Extract data for Excel
    core_dependencies = table_data['core_dependencies']
    other_dependencies = table_data['other_dependencies']
    all_dependencies = {**other_dependencies, **core_dependencies}  # shouldn't be overlap, but core values would win
    project_data = table_data['project_data']

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # Write headers to the worksheet
    worksheet['A1'] = 'Package'
    worksheet['B1'] = 'Allowed Version Range'

    # Write data to the worksheet
    if core_dependencies:
        # Write core environment dependencies to the first two columns of the worksheet (shaded light gray)
        for package_name, (row, version_range, spec0_compliant) in core_dependencies.items():
            # Column 1 - Package
            cell = worksheet.cell(row=row, column=1)
            cell.value = package_name
            cell.fill = GRAY

            # Column 2 - Allowed Version Range (with SPEC 0 highlighting if applicable)
            cell = worksheet.cell(row=row, column=2)
            cell.value = version_range  # TODO: shouldn't ever be "N/A" but should I consider that case here anyway?
            # Keep gray background for core dependencies, but note SPEC 0 status is tracked
            cell.fill = GRAY

        # Add a dark gray separator before other dependencies
        cell = worksheet.cell(row=row+1, column=1)
        cell.fill = DARK_GRAY
        cell = worksheet.cell(row=row+1, column=2)
        cell.fill = DARK_GRAY

    # Write other environment dependencies to the first two columns of the worksheet
    for package_name, (row, version_range, spec0_compliant) in other_dependencies.items():
        # Column 1 - Package
        cell = worksheet.cell(row=row, column=1)
        cell.value = package_name

        # Column 2 - Allowed Version Range (with SPEC 0 highlighting)
        cell = worksheet.cell(row=row, column=2)
        if version_range is not None:
            cell.value = version_range
            # Yellow if SPEC 0 non-compliant, no color if compliant
            if spec0_compliant is False:
                cell.fill = YELLOW
        else:
            # "N/A" means dependency conflict
            cell.value = "N/A"
            # Always red for conflicts in the allowed range column
            cell.fill = RED

    # Write project data to the worksheet one column at a time
    for j, (project, dependency_data) in enumerate(project_data.items(), start=3):
        worksheet.cell(row=1, column=j, value=project)  # column header
        for package_name, (compatible, spec0_compliant, version_range) in dependency_data.items():
            row_num = all_dependencies[package_name][0]
            allowed_spec0_compliant = all_dependencies[package_name][2]  # SPEC 0 status of "Allowed Version Range"
            cell = worksheet.cell(row=row_num, column=j)
            if compatible is not None:
                cell.value = version_range

                # Determine cell color based on compatibility and SPEC 0 compliance
                if compatible:
                    # Green or Yellow: Package is compatible with allowed range
                    # Only check SPEC 0 if the "Allowed Version Range" is NOT definitively compliant
                    if allowed_spec0_compliant is not True and spec0_compliant is False:
                        # Yellow: Compatible but SPEC 0 non-compliant
                        cell.fill = YELLOW
                    else:
                        # Green: Compatible (and either SPEC 0 compliant or not a SPEC 0 package)
                        cell.fill = GREEN
                else:
                    # Red or Orange: Package is incompatible with allowed range
                    # Only check SPEC 0 if the "Allowed Version Range" is NOT definitively compliant
                    if allowed_spec0_compliant is not True and spec0_compliant is False:
                        # Orange: Incompatible AND SPEC 0 non-compliant
                        cell.fill = ORANGE
                    else:
                        # Red: Incompatible (regardless of SPEC 0 status)
                        cell.fill = RED

    return workbook


if __name__ == '__main__':
    # Legacy direct script mode is intentionally disabled for now.
    # The workflow uses pipeline_v2.py --generate-spreadsheet.
    pass
